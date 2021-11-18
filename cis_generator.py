import argparse, re, pdb
from string import ascii_letters
from ssp import SSP
from openpyxl import load_workbook

class CIS_Control:

    implementation_columns = {'Implemented': 1, 'Partially Implemented': 2, 'Planned': 3, 'Alternative Implementation': 4, 'Not Applicable': 5}
    origination_columns = {
        'Service Provider Corporate': 6,
        'Service Provider System Specific': 7,
        'Service Provider Hybrid': 8,
        'Configured by Customer': 9,
        'Provided by Customer': 10,
        'Shared': 11,
        'Inherited': 12,
        'Not Applicable': 13
        }

    def __init__(self, control_object):
        self.number = control_object.number
        self.implementation_status = control_object.implementation_status
        self.control_origination = control_object.control_origination

    def __repr__(self):
        return self.number

    def get_columns(self):
        relevant_columns = []
        try:
            for status in self.implementation_status:
                relevant_columns.append(self.implementation_columns[status])
            for origin in self.control_origination:
                relevant_columns.append(self.origination_columns[origin])
        except KeyError:
            pass
        return relevant_columns


class CRM_Control:
    def __init__(self, number, text):
        self.number = number
        self.text = text
    
    def __repr__(self):
        return self.number

def get_control_parts(control_object):
    parts = []
    for part in control_object:
        if part is None:
            cust_resp = get_customer_responsibility_text(control_object.part(None).text)
            if cust_resp:
                part_obj = CRM_Control(control_object.number, cust_resp)
                parts.append(part_obj)
        else:
            cust_resp = get_customer_responsibility_text(control_object.part(part).text)
            if cust_resp:
                part_num = create_part_num(control_object.number, part)
                part_obj = CRM_Control(part_num, cust_resp)
                parts.append(part_obj)
    return parts

def create_part_num(control_number, part_number):
    part_num = "%s(%s)" % (control_number, part_number)
    return part_num


def get_customer_responsibility_text(control_text):
    if "Customer Responsibility" in control_text:
        cust_resp = ''
        split_text = control_text.split('\n')
        for text in split_text:
            if "Customer Resp" in text:
                customer_index = split_text.index(text) 
        for text_part in split_text[customer_index:]:
            if 'Customer Responsibility' in text_part:
                continue
            if ':' in text_part and "Part" in text_part and "http" not in text_part:
                return cust_resp
            elif ':' in text_part[-3:] and 'http' not in text_part and 'as:' not in text_part and 'link:' not in text_part:
                return cust_resp
            else:
                cust_resp = cust_resp + text_part
        return cust_resp
    else:
        return None

def get_max_row(worksheet):
    row = 4
    value = worksheet.cell(row, 1).value
    while value is not None:
        row+=1
        value = worksheet.cell(row, 1).value
    return row


def fill_cis_worksheet(cis_dict, worksheet):
    rownumber = 4
    max_row = get_max_row(worksheet)
    while rownumber < max_row:
        row = worksheet[rownumber]
        control = row[0].value
        control = convert_cis_control_number(control)
        try:
            control_object = cis_dict[control]
        except KeyError:
            print('Could not find entry for control ' + control)
            rownumber += 1
            continue
        columns = control_object.get_columns()
        for column in columns:
            row[column].value = 'X'
        rownumber += 1

def create_addendum_controls(addendum):
    crm_addendum_list = []
    cis_addendum_list = []
    for control in addendum:
        cis_addendum_list.append(CIS_Control(control))
        crm_addendum_list.extend(get_addendum_control_parts(control.implementation_table, control))
    return cis_addendum_list, crm_addendum_list

def get_addendum_control_parts(table, control_object):
    parts = []
    for part in control_object.parts:
        if part is None:
            cust_resp = get_customer_responsibility_text(table.cell(1,0).text)
            if cust_resp:
                part_obj = CRM_Control(control_object.number, cust_resp)
                parts.append(part_obj)

        else:
            for row in table.rows:
                if 'Part ' + part.strip() in row.cells[0].text.strip():
                    cust_resp = get_customer_responsibility_text(row.cells[1].text.strip())
                    if cust_resp:
                        part_num = create_part_num(control_object.number, part)
                        part_obj = CRM_Control(part_num, cust_resp)
                        parts.append(part_obj)
    return parts

def append_addendum_controls_to_cis(control_list, cis_worksheet):
    rows = []
    for control in control_list:
        new_row = [''] * 15
        new_row[1] = control.number
        columns = control.get_columns()
        for column in columns:
            new_row[column] = 'X'
        cis_worksheet.append(new_row)

# def append_addendum_controls_to_crm(control_list, crm_worksheet):
#     rows = []
#     for control in control_list:

def main(docs, cis_workbook, out_file):
    cis_worksheet = cis_workbook['CIS Worksheet']
    crm_worksheet = cis_workbook['CRM Worksheet']
    security_plan, addendum = docs
    crm_control_list = []
    cis_control_dict = {}
    for control in security_plan:
        # print(control.number)
        cis_control_dict[control.number] = CIS_Control(control)
        crm_control_list.extend(get_control_parts(control))
    if addendum:
        cis_addendum_list, crm_addendum_list = create_addendum_controls(addendum)
    else:
        crm_addendum_list = None
    
    cis_controls = [convert_cis_control_number(cis_worksheet.cell(row=x, column = 1).value) for x in range(4, get_max_row(cis_worksheet))]
    for control in [control for control in security_plan if control.number not in cis_controls]:
        new_row = [''] * 15
        new_row[1] = control.number
        cis_worksheet.append(new_row)
    fill_cis_worksheet(cis_control_dict, cis_worksheet)
    fill_crm_worksheet(crm_control_list, crm_worksheet, crm_addendum_list)
    if addendum:
        append_addendum_controls_to_cis(cis_addendum_list, cis_worksheet)
        # append_addendum_controls_to_crm(crm_addendum_list, crm_worksheet)
    cis_workbook.save(out_file)

def fill_crm_worksheet(crm_control_list, crm_worksheet, crm_addendum_list):
    crm_worksheet_dict = {}
    for row in range(4, get_max_row(crm_worksheet)):
        cisnumber = crm_worksheet.cell(row, 1).value
        crm_worksheet.cell(row=row, column=2).value = "Yes"
        if cisnumber is None:
            continue
        if any(ascii in cisnumber for ascii in ascii_letters):
            cisnumber = re.sub(r' (?=\([a-z]\))', '', cisnumber)
        if '0' in cisnumber:
            cisnumber = re.sub(r'0(?=[1-9])', '', cisnumber)
        crm_worksheet_dict[cisnumber] = row
    for control in crm_control_list:
        row_counter = crm_worksheet_dict[control.number]
        #crm_worksheet.cell(row_counter, 1).value = ref_counter
        crm_worksheet.cell(row_counter, 1).value = control.number
        crm_worksheet.cell(row_counter, 2).value = 'No'
        crm_worksheet.cell(row_counter, 3).value = control.text
    if crm_addendum_list:
        for control in crm_addendum_list:
            row_counter = crm_worksheet_dict[control.number]
            #crm_worksheet.cell(row_counter, 1).value = ref_counter
            crm_worksheet.cell(row_counter, 1).value = control.number
            crm_worksheet.cell(row_counter, 2).value = 'Yes'
            crm_worksheet.cell(row_counter, 3).value = control.text


def convert_cis_control_number(control_number):
    control_number = control_number.replace('-0', '-').replace('(0', '(').split(' ', 1)[0]
    return control_number


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--ssp', help='Path to SSP', required=True)
    parser.add_argument('--template', help='Path to CIS template', required=True)
    parser.add_argument('--out', help='name of output file, include .xlsx', required=True)
    parser.add_argument('--addendum', help='Path to addendum if applicable', required=False)
    args = parser.parse_args()
    
    security_plan = SSP(args.ssp)
    cis_workbook = load_workbook(args.template)
    result_file = args.out
    if args.addendum:
        addendum = SSP(args.addendum)
    else:
        addendum = None
    main([security_plan, addendum], cis_workbook, result_file)
